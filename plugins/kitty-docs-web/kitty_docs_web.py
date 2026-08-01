# /// script
# dependencies = [
#   "fastmcp",
#   "httpx",
#   "trafilatura",
#   "ddgs",
#   "openpyxl",
#   "pymupdf",
# ]
# ///
"""kitty-docs-web: PDF/Excel reading, web scraping, and web search.

Split out of `plugins/replacement-mcp/lean_mcp.py` — see that file's
top-of-file note. Tool names are kept byte-identical to their old
`lean_mcp.py` registrations (adaptive-pathway keys learned routing
preferences on the literal name string; renaming orphans that history — see
`docs/PLUGINS.md`).

`success_response`/`error_response`/`_filter_by_query` are duplicated from
`lean_mcp.py` rather than imported (two separate PyInstaller bundles; a
shared package would cost more than it saves) — kept behaviorally identical
except where a fix below is explicitly called out.

`lean_web_search` merges what used to be two separate tools —
`brave_mcp_search` (Rust, `kitty-tools`) and `lean_fallback_web_search`
(this file) — into one count-tiered tool: Brave-with-DuckDuckGo-fallback for
small requests, both engines queried together for broader requests, and an
offloaded/indexed mode for large ones. See `docs/VERSIONS.md` for why it
lives here rather than in Rust: DuckDuckGo's `ddgs` has no Rust equivalent,
while Brave's call is a plain JSON GET, trivial to host alongside it.
"""

import concurrent.futures
import csv
import io
import json
import os
import random
import re
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import fitz  # PyMuPDF
import httpx
from fastmcp import FastMCP

mcp = FastMCP("kitty-docs-web")

# Same cache directory `lean_mcp.py`'s cache tools already manage — a
# downloaded PDF here is meant to be immediately readable by
# `lean_pdf_read_text`/`lean_pdf_read_outline`, which live in this same
# process, and inspectable via `lean_cache_view` (still hosted by
# replacement-mcp) without needing to agree on a second cache directory.
CACHE_DIR = Path.home() / ".cache" / "lean-goose-mcp"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# Plain module constants rather than a YAML config layer — this plugin is
# small enough that a second `tool_prompts.yaml`-style file would be pure
# overhead. Same default value as the old `scrape_max_chars` threshold.
SCRAPE_MAX_CHARS_DEFAULT = 12000
EXCEL_MAX_ROWS_DEFAULT = 500


# ---------------------------------------------------------------------------
# Standardized JSON response helpers (kept byte-identical to lean_mcp.py's)
# ---------------------------------------------------------------------------
def success_response(
    data: Any,
    message: Optional[str] = None,
    truncated: bool = False,
    metadata: Optional[Dict[str, Any]] = None,
) -> str:
    """Returns a standardized JSON success response."""
    payload: Dict[str, Any] = {
        "status": "success",
        "truncated": truncated,
        "data": data,
    }
    if message:
        payload["message"] = message
    if metadata:
        payload["metadata"] = metadata
    return json.dumps(payload, indent=2, ensure_ascii=False)


def error_response(
    code: str,
    message: str,
    detail: Optional[str] = None,
    hint: Optional[str] = None,
) -> str:
    """Returns a standardized JSON error response with automated recovery hints."""
    payload: Dict[str, Any] = {
        "status": "error",
        "error_code": code,
        "message": message,
    }
    if detail:
        payload["detail"] = detail

    if not hint:
        if "NOT_FOUND" in code or "MISSING" in code:
            hint = "Verify path spelling or call lean_analyze_workspace to check available files."
        elif "CORRUPT" in code or "PARSE" in code:
            hint = "File may be damaged or password-protected. Verify format."
        elif "BAD_RANGE" in code or "OUT_OF_BOUNDS" in code:
            hint = "Inspect dimensions or line counts before specifying bounds."
        elif "SEARCH" in code:
            hint = "Broaden search keywords or check network connectivity."
        elif "SCRAPE" in code:
            # Track C fix: the old shared hint mapper matched "SEARCH or
            # SCRAPE" on one branch, so every scrape failure got search
            # advice ("Broaden search keywords..."). Every scrape error code
            # below sets its own explicit hint, so this is just a backstop
            # for an unanticipated one.
            hint = "Try a different URL, or use lean_web_search to find an alternative source."

    if hint:
        payload["hint"] = hint
    return json.dumps(payload, indent=2, ensure_ascii=False)


# ---------------------------------------------------------------------------
# In-tool keyword RAG helper (Track E fix: offset-based continuation, no
# fabricated string injected into `data` on a no-match)
# ---------------------------------------------------------------------------
@dataclass
class QueryFilterResult:
    items: List[str]
    truncated: bool
    total_matches: int
    next_offset: Optional[int]
    # True when a query was given but nothing scored > 0 — the caller should
    # surface this as `message`, not fabricate a line inside `data`.
    no_match: bool


def _filter_by_query(
    items: List[str],
    query: Optional[str] = None,
    max_results: int = 50,
    offset: int = 0,
) -> QueryFilterResult:
    """Filters lines/paragraphs/rows by keyword match score.

    Stable descending sort — ties keep document order (`scored` is built in
    index order and Python's `list.sort` is stable). Deliberately NOT
    `sort_by_key(...); reverse()` — that flips the order of tied items.
    """
    if not query or not query.strip():
        total = len(items)
        page = items[offset : offset + max_results]
        has_more = offset + len(page) < total
        return QueryFilterResult(page, has_more, total, offset + len(page) if has_more else None, False)

    query_words = set(re.findall(r"\w+", query.lower()))
    if not query_words:
        total = len(items)
        page = items[offset : offset + max_results]
        has_more = offset + len(page) < total
        return QueryFilterResult(page, has_more, total, offset + len(page) if has_more else None, False)

    scored = []
    for idx, item in enumerate(items):
        item_words = set(re.findall(r"\w+", item.lower()))
        score = len(query_words.intersection(item_words))
        if score > 0:
            scored.append((score, idx, item))

    if not scored:
        page = items[offset : offset + max_results]
        has_more = offset + len(page) < len(items)
        return QueryFilterResult(page, has_more, 0, offset + len(page) if has_more else None, True)

    scored.sort(key=lambda x: x[0], reverse=True)
    total_matches = len(scored)
    page = [item for _, _, item in scored[offset : offset + max_results]]
    has_more = offset + len(page) < total_matches
    return QueryFilterResult(page, has_more, total_matches, offset + len(page) if has_more else None, False)


# ===========================================================================
# Web & Search Tools (Track C redesign)
# ===========================================================================
_TRACKING_PARAM_PREFIXES = ("utm_",)
_TRACKING_PARAM_EXACT = {"fbclid", "gclid", "msclkid", "mc_cid", "mc_eid", "ref_src", "igshid"}


def _strip_tracking_params(url: str) -> str:
    """Removes known tracking/analytics query params while preserving any
    parameter that's part of the actual resource address (e.g. YouTube's
    `v=`, a wiki's `?id=`, a forum's `?p=`).

    The old code did `re.sub(r"\\?.*$", "", href)` — a blanket strip of the
    entire query string. For any URL where the query string *is* the
    resource, that "cleaned" URL points somewhere else or 404s, and a model
    that then hands it to `lean_web_scrape` gets a broken-link error it can't
    explain. An allowlist of known tracking params fixes the search→scrape
    chain without losing real resource identifiers.
    """
    if not url or "?" not in url:
        return url
    parsed = urlparse(url)
    kept = [
        (k, v)
        for k, v in parse_qsl(parsed.query, keep_blank_values=True)
        if not (k.lower().startswith(_TRACKING_PARAM_PREFIXES) or k.lower() in _TRACKING_PARAM_EXACT)
    ]
    return urlunparse(parsed._replace(query=urlencode(kept)))


# ---------------------------------------------------------------------------
# lean_web_search: count-tiered Brave/DuckDuckGo merge.
#
# count <= NORMAL_MAX_COUNT (default 5): Brave first if configured, DuckDuckGo
#   only as a fallback on Brave failure. Inline, full detail.
# NORMAL_MAX_COUNT < count <= EXPANDED_MAX_COUNT: Brave AND DuckDuckGo queried
#   CONCURRENTLY regardless of whether Brave succeeds — "expansion" means
#   broadening sources, not inventing lexical query variants (a deterministic
#   synonym/thesaurus step would need either an LLM or a maintenance-heavy
#   static table; querying both engines is cheap and needs neither). Still
#   inline, full detail.
# count > EXPANDED_MAX_COUNT: same dual-engine fetch, but the full result set
#   is written to a temp offload file and a compact keyword-based index is
#   returned instead (no ranking — see _extract_keywords). Follow up with
#   lean_web_search_read_chunk for full detail on chosen ids.
# ---------------------------------------------------------------------------
BRAVE_API_KEY = os.environ.get("BRAVE_API_KEY", "")
# Sibling to CACHE_DIR, not inside it — so a future cache-clear tool can never
# touch an in-flight search offload (mirrors the Rust side's scratchpad
# relocation for the identical reason).
SEARCH_STORE_DIR = Path.home() / ".cache" / "kitty-search-offload"
MAX_OFFLOAD_FILES = 20
MAX_RATE_LIMIT_RETRIES = 2
BASE_BACKOFF_SECONDS = 1.5

NORMAL_MAX_COUNT = 5
EXPANDED_MAX_COUNT = 10
MAX_COUNT = 50
KEYWORDS_PER_ITEM = 5
MANIFEST_TITLE_MAX_CHARS = 60

_QUERY_PREFIX_RE = re.compile(
    r"^(please\s+)?(search\s+for|find\s+me|look\s+up|what\s+is|who\s+is|where\s+is|can\s+you\s+search\s+for)\s+",
    re.IGNORECASE,
)


def _apply_query_guardrails(query: str) -> str:
    """Strips conversational fluff and clamps length before hitting an API."""
    q = query.strip().strip("\"'")
    q = _QUERY_PREFIX_RE.sub("", q).strip()
    words = q.split()
    if len(words) > 50:
        q = " ".join(words[:50])
    return q[:400]


def _normalize_url_key(url: str) -> str:
    """Lowercased, tracking-stripped URL used only for dedup matching (never
    shown to the caller)."""
    cleaned = _strip_tracking_params(url).lower()
    if cleaned.startswith("https://"):
        cleaned = cleaned[len("https://") :]
    elif cleaned.startswith("http://"):
        cleaned = cleaned[len("http://") :]
    if cleaned.startswith("www."):
        cleaned = cleaned[len("www.") :]
    return cleaned.rstrip("/")


class _BraveFailure(Exception):
    """kind in {"rate_limit_exhausted", "auth", "network", "api", "invalid_query"}.
    "invalid_query" (HTTP 400/422) is a caller error, not an availability
    failure — callers must NOT treat it as a fallback trigger."""

    def __init__(self, kind: str, detail: str = ""):
        self.kind = kind
        self.detail = detail
        super().__init__(f"{kind}: {detail}")


def _brave_query(
    query: str, count: int, search_lang: str, freshness: Optional[str], country: str
) -> List[Dict[str, Any]]:
    """Calls Brave's LLM-context search API with bounded 429 retry/backoff."""
    params: Dict[str, Any] = {
        "q": query,
        "count": max(1, min(count, 50)),
        "search_lang": search_lang,
        "country": country,
    }
    if freshness:
        params["freshness"] = freshness

    for attempt in range(MAX_RATE_LIMIT_RETRIES + 1):
        try:
            response = httpx.get(
                "https://api.search.brave.com/res/v1/llm/context",
                params=params,
                headers={
                    "X-Subscription-Token": BRAVE_API_KEY,
                    "Accept": "application/json",
                    "Accept-Encoding": "gzip",
                },
                timeout=httpx.Timeout(connect=10.0, read=30.0, write=10.0, pool=10.0),
            )
        except httpx.RequestError as e:
            raise _BraveFailure("network", str(e))

        if response.status_code == 429:
            if attempt < MAX_RATE_LIMIT_RETRIES:
                retry_after = response.headers.get("Retry-After")
                delay = None
                if retry_after is not None:
                    try:
                        delay = max(0.0, float(retry_after))
                    except ValueError:
                        delay = None
                if delay is None:
                    delay = BASE_BACKOFF_SECONDS * (2**attempt) + random.uniform(0, 0.5)
                time.sleep(delay)
                continue
            raise _BraveFailure("rate_limit_exhausted", response.text)

        if response.status_code in (400, 422):
            raise _BraveFailure("invalid_query", response.text)
        if response.status_code in (401, 403):
            raise _BraveFailure("auth", response.text)
        if not response.is_success:
            raise _BraveFailure("api", f"HTTP {response.status_code}: {response.text}")

        try:
            payload = response.json()
        except Exception as e:
            raise _BraveFailure("api", f"unparseable response: {e}")
        return _parse_brave_results(payload)

    raise _BraveFailure("rate_limit_exhausted", "retries exhausted")


def _parse_brave_results(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extracts a flat result list from Brave's grounding/sources response shape."""
    grounding = payload.get("grounding") or {}
    sources = {s.get("url", ""): s for s in (payload.get("sources") or [])}

    def _clean(raw_url: str) -> tuple[str, str]:
        clean_url = _strip_tracking_params(raw_url)
        return clean_url, urlparse(clean_url).netloc

    items: List[Dict[str, Any]] = []
    for entry in grounding.get("generic") or []:
        raw_url = entry.get("url", "")
        clean_url, domain = _clean(raw_url)
        source = sources.get(raw_url, {})
        snippets = entry.get("snippets") or ([entry["snippet"]] if entry.get("snippet") else [])
        items.append(
            {
                "title": entry.get("title", ""),
                "domain": domain or source.get("hostname", ""),
                "url": clean_url,
                "date": entry.get("age") or source.get("date"),
                "snippet": " ".join(snippets).strip(),
            }
        )

    poi = grounding.get("poi")
    if poi:
        raw_url = poi.get("url", "")
        clean_url, domain = _clean(raw_url)
        items.append(
            {
                "title": poi.get("title", ""),
                "domain": domain,
                "url": clean_url,
                "date": None,
                "snippet": poi.get("description", ""),
            }
        )

    for entry in grounding.get("map") or []:
        raw_url = entry.get("url", "")
        clean_url, domain = _clean(raw_url)
        items.append(
            {
                "title": entry.get("title", ""),
                "domain": domain,
                "url": clean_url,
                "date": None,
                "snippet": entry.get("description", ""),
            }
        )

    return items


def _ddg_query(query: str, count: int) -> List[Dict[str, Any]]:
    """Single-shot DuckDuckGo search, no retry — unchanged from the retired
    lean_fallback_web_search's behavior."""
    from ddgs import DDGS

    capped = max(1, min(count, 20))
    raw_results = list(DDGS().text(query, max_results=capped))
    cleaned = []
    for r in raw_results:
        clean_url = _strip_tracking_params(r.get("href", ""))
        cleaned.append(
            {
                "title": r.get("title", ""),
                "domain": urlparse(clean_url).netloc,
                "url": clean_url,
                "date": None,
                "snippet": r.get("body", ""),
            }
        )
    return cleaned


def _normal_search(
    query: str, count: int, search_lang: str, freshness: Optional[str], country: str
) -> tuple[List[Dict[str, Any]], Dict[str, str]]:
    """count <= NORMAL_MAX_COUNT: Brave-first-if-configured, DuckDuckGo only as a
    failure fallback. `invalid_query` is a caller error and propagates directly
    rather than triggering a fallback."""
    diagnostics = {"brave": "not_configured", "duckduckgo": "not_queried"}
    if BRAVE_API_KEY:
        try:
            results = _brave_query(query, count, search_lang, freshness, country)
            diagnostics["brave"] = "ok"
            for item in results:
                item["engine"] = "brave"
            return results, diagnostics
        except _BraveFailure as e:
            if e.kind == "invalid_query":
                raise
            diagnostics["brave"] = e.kind

    diagnostics["duckduckgo"] = "ok"
    results = _ddg_query(query, count)
    for item in results:
        item["engine"] = "duckduckgo"
    return results, diagnostics


def _dual_engine_search(
    query: str, count: int, search_lang: str, freshness: Optional[str], country: str
) -> tuple[List[Dict[str, Any]], Dict[str, str]]:
    """count > NORMAL_MAX_COUNT: queries Brave AND DuckDuckGo concurrently
    regardless of whether Brave succeeds — this is the "expansion": broader
    source coverage, not lexical query variants. Brave's `invalid_query` does
    not abort DuckDuckGo's side; it's only recorded in diagnostics."""
    diagnostics: Dict[str, str] = {"brave": "not_configured", "duckduckgo": "not_queried"}
    brave_results: List[Dict[str, Any]] = []
    ddg_results: List[Dict[str, Any]] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
        ddg_future = pool.submit(_ddg_query, query, count)
        brave_future = None
        if BRAVE_API_KEY:
            brave_future = pool.submit(_brave_query, query, count, search_lang, freshness, country)

        if brave_future is not None:
            try:
                brave_results = brave_future.result()
                diagnostics["brave"] = "ok"
            except _BraveFailure as e:
                diagnostics["brave"] = e.kind

        try:
            ddg_results = ddg_future.result()
            diagnostics["duckduckgo"] = "ok"
        except Exception as e:
            diagnostics["duckduckgo"] = f"failed: {e}"

    for item in brave_results:
        item["engine"] = "brave"
    for item in ddg_results:
        item["engine"] = "duckduckgo"

    merged = list(brave_results)
    seen = {_normalize_url_key(r["url"]) for r in brave_results}
    for item in ddg_results:
        key = _normalize_url_key(item["url"])
        if key not in seen:
            seen.add(key)
            merged.append(item)

    return merged, diagnostics


# ~150 common English stopwords, filtered out before frequency-counting a
# result's title+snippet. Frequency-based (TF), not LSA: at this scale (a
# handful of short blurbs) there isn't enough text for SVD to find real
# latent structure beyond what plain term frequency already shows, and
# scikit-learn/numpy would be a new PyInstaller dependency in a project whose
# whole rewrite motivation was shrinking frozen binaries.
_STOPWORDS = frozenset(
    """
    a about above after again against all am an and any are aren't as at be because been
    before being below between both but by can't cannot could couldn't did didn't do does
    doesn't doing don't down during each few for from further had hadn't has hasn't have
    haven't having he he'd he'll he's her here here's hers herself him himself his how
    how's i i'd i'll i'm i've if in into is isn't it it's its itself let's me more most
    mustn't my myself no nor not of off on once only or other ought our ours ourselves out
    over own same shan't she she'd she'll she's should shouldn't so some such than that
    that's the their theirs them themselves then there there's these they they'd they'll
    they're they've this those through to too under until up very was wasn't we we'd we'll
    we're we've were weren't what what's when when's where where's which while who who's
    whom why why's will with won't would wouldn't you you'd you'll you're you've your
    yours yourself yourselves new using use used via also into
    """.split()
)

_WORD_RE = re.compile(r"[a-zA-Z][a-zA-Z\-']+")


def _extract_keywords(text: str, top_k: int = KEYWORDS_PER_ITEM) -> List[str]:
    """Deterministic, frequency-based keyword extraction. `Counter` preserves
    first-insertion order and Python's sort is stable, so sorting only by
    `-count` keeps first-appearance order as the tie-break — the same idiom
    `_filter_by_query` above already relies on for reproducible ties."""
    words = _WORD_RE.findall(text.lower())
    counts = Counter(w for w in words if w not in _STOPWORDS and len(w) > 2)
    ranked = sorted(counts.items(), key=lambda kv: -kv[1])
    return [w for w, _ in ranked[:top_k]]


def _build_index(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """No ranking/scoring across items — order is just merge order
    (Brave-first, then DuckDuckGo-only). `keywords` is the deterministic
    per-item descriptor; `url`/`snippet`/`date` are deferred entirely to
    lean_web_search_read_chunk."""
    return [
        {
            "id": r["id"],
            "title": r["title"][:MANIFEST_TITLE_MAX_CHARS],
            "domain": r["domain"],
            "engine": r["engine"],
            "keywords": _extract_keywords(f"{r['title']} {r['snippet']}"),
        }
        for r in results
    ]


def _search_id() -> str:
    return f"{int(time.time() * 1000):x}-{random.randint(0, 0xffff):04x}"


def _offload_path(search_id: str) -> Path:
    return SEARCH_STORE_DIR / f"search-{search_id}.json"


def _prune_old_offloads() -> None:
    if not SEARCH_STORE_DIR.exists():
        return
    files = sorted(
        SEARCH_STORE_DIR.glob("search-*.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    # -1: room for the new file about to be written.
    for stale in files[MAX_OFFLOAD_FILES - 1 :]:
        try:
            stale.unlink()
        except OSError:
            pass


def _write_offload(search_id: str, query: str, results: List[Dict[str, Any]]) -> None:
    SEARCH_STORE_DIR.mkdir(parents=True, exist_ok=True)
    _prune_old_offloads()
    _offload_path(search_id).write_text(
        json.dumps({"search_id": search_id, "query": query, "results": results}, ensure_ascii=False),
        encoding="utf-8",
    )


@mcp.tool(name="lean_web_search")
def web_search(
    query: str,
    count: int = 5,
    search_lang: str = "en",
    freshness: Optional[str] = None,
    country: str = "US",
) -> str:
    """Searches the web. count<=5 (default): Brave if configured, DuckDuckGo
    only as a fallback on Brave failure. count 6-10: queries Brave AND
    DuckDuckGo together for broader coverage, still returned inline.
    count>10: same broadened fetch, but the full result set is offloaded to
    disk and a compact keyword index is returned instead of full detail —
    follow up with lean_web_search_read_chunk for full detail on chosen ids.
    """
    count = max(1, min(count, MAX_COUNT))
    guarded_q = _apply_query_guardrails(query)
    if not guarded_q:
        return error_response(
            "EMPTY_QUERY", "The search query was empty or contained only whitespace."
        )

    if count <= NORMAL_MAX_COUNT:
        mode = "normal"
    elif count <= EXPANDED_MAX_COUNT:
        mode = "expanded"
    else:
        mode = "expansive"

    try:
        if mode == "normal":
            results, diagnostics = _normal_search(guarded_q, count, search_lang, freshness, country)
        else:
            results, diagnostics = _dual_engine_search(guarded_q, count, search_lang, freshness, country)
    except _BraveFailure as e:
        # Only reachable for "invalid_query" — every other kind is swallowed
        # into a fallback attempt inside _normal_search/_dual_engine_search.
        return error_response(
            "INVALID_QUERY",
            "Brave Search API rejected the query parameters.",
            detail=e.detail,
            hint="Simplify the query text or check freshness/country values.",
        )

    if not results:
        return error_response(
            "NO_RESULTS",
            "No results found across the configured search engines.",
            hint="Try a broader or different query, or increase count.",
        )

    if mode in ("normal", "expanded"):
        return success_response(
            results[:count],
            metadata={"mode": mode, "engines": diagnostics, "query": guarded_q, "count": count},
        )

    for idx, item in enumerate(results, start=1):
        item["id"] = idx
    search_id = _search_id()
    _write_offload(search_id, guarded_q, results)
    manifest = _build_index(results[:count])
    return success_response(
        manifest,
        metadata={
            "mode": mode,
            "engines": diagnostics,
            "search_id": search_id,
            "query": guarded_q,
            "count": count,
            "total_results_found": len(results),
        },
    )


@mcp.tool(name="lean_web_search_read_chunk")
def web_search_read_chunk(search_id: str, ids: List[int]) -> str:
    """Fetches full url/snippet/date detail for specific result ids from a
    prior lean_web_search expansive-mode index."""
    if "/" in search_id or "\\" in search_id or ".." in search_id:
        return error_response(
            "SEARCH_ID_NOT_FOUND", "Invalid search_id.", hint="Call lean_web_search again."
        )

    path = _offload_path(search_id)
    if not path.exists():
        return error_response(
            "SEARCH_ID_NOT_FOUND",
            f"No stored search results for search_id '{search_id}'.",
            hint="This search_id may have expired (only the 20 most recent searches are "
            "retained) or been mistyped; call lean_web_search again.",
        )

    try:
        stored = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return error_response(
            "SEARCH_ID_NOT_FOUND",
            f"Stored search results are corrupt: {e}",
            hint="Call lean_web_search again.",
        )

    requested = ids[:5]
    truncated_ids = len(ids) > 5
    by_id = {r["id"]: r for r in stored.get("results", [])}
    matched = [by_id[i] for i in requested if i in by_id]

    if not matched:
        return error_response(
            "ID_NOT_FOUND",
            "None of the requested ids exist in this search.",
            hint="Ids come from a lean_web_search expansive-mode response; call it again if unsure.",
        )

    meta: Dict[str, Any] = {"search_id": search_id}
    if truncated_ids:
        meta["ids_truncated_to"] = 5
    return success_response(matched, metadata=meta)


# A complete, current-looking UA string. The old one ended mid-string at
# "AppleWebKit/537.36" with no "(KHTML, like Gecko) Chrome/... Safari/..."
# tail — a shape several WAFs fingerprint directly as non-browser traffic,
# and a straightforward, avoidable cause of 403s.
_SCRAPE_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)


def _split_markdown_blocks(text: str) -> List[str]:
    """Splits markdown on blank-line boundaries, keeping a fenced code block
    intact even if it contains a blank line (a naive '\\n\\n'.split would
    otherwise slice through an open ``` fence and reassemble it as two
    separate, malformed "paragraphs")."""
    normalized = re.sub(r"\n{3,}", "\n\n", text.strip())
    raw_blocks = normalized.split("\n\n")
    blocks: List[str] = []
    buffer: List[str] = []
    fence_open = False
    for block in raw_blocks:
        buffer.append(block)
        if block.count("```") % 2 == 1:
            fence_open = not fence_open
        if not fence_open:
            blocks.append("\n\n".join(buffer))
            buffer = []
    if buffer:
        blocks.append("\n\n".join(buffer))
    return [b.strip() for b in blocks if b.strip()]


def _strip_markdown_links(text: str) -> str:
    """[label](url) -> label, but leaves fenced code blocks untouched so a
    URL inside a code sample is never rewritten."""
    parts = re.split(r"(```.*?```)", text, flags=re.DOTALL)
    out = []
    for i, part in enumerate(parts):
        if i % 2 == 1:
            out.append(part)
        else:
            out.append(re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", part))
    return "".join(out)


def _cap_blocks_by_chars(blocks: List[str], cap: int) -> tuple[str, int]:
    """Greedily takes whole blocks up to `cap` characters, never cutting
    mid-block (mid-table, mid-heading) except when a single block alone
    already exceeds the cap, in which case that one block is truncated so
    the tool still returns *something* rather than nothing."""
    if not blocks:
        return "", 0
    if len(blocks[0]) > cap:
        return blocks[0][:cap], 1
    selected: List[str] = []
    running = 0
    for b in blocks:
        block_len = len(b) + (2 if selected else 0)
        if selected and running + block_len > cap:
            break
        selected.append(b)
        running += block_len
    return "\n\n".join(selected), len(selected)


@mcp.tool(name="lean_web_scrape")
def web_scrape(
    url: str,
    query: Optional[str] = None,
    output_format: Literal["markdown", "text"] = "markdown",
    offset: int = 0,
    max_chars: Optional[int] = None,
    include_links: bool = False,
    favor_precision: bool = False,
) -> str:
    """Scrapes a URL into clean Markdown (or plain text) for an LLM to read.

    `offset` and the response's `metadata.next_offset` page through long
    pages by markdown block, without severing a table or heading mid-cut.
    A PDF URL is downloaded to the cache and its local path is returned —
    use `lean_pdf_read_text`/`lean_pdf_read_outline` on that path next.
    `favor_precision=True` favors precision over recall (trafilatura's
    aggressive extraction mode); the default favors recall, since
    documentation/API-reference pages often lose sidebars and short
    definition blocks under the precision-favoring mode.
    """
    stripped_url = url.split("?")[0].split("#")[0]
    looks_like_pdf = stripped_url.lower().endswith(".pdf")

    try:
        response = httpx.get(
            url,
            timeout=httpx.Timeout(connect=10.0, read=30.0, write=10.0, pool=10.0),
            follow_redirects=True,
            headers={
                "User-Agent": _SCRAPE_USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf;q=0.8,*/*;q=0.5",
                "Accept-Language": "en-US,en;q=0.9",
            },
        )
        response.raise_for_status()
    except httpx.HTTPStatusError as e:
        status = e.response.status_code
        return error_response(
            "SCRAPE_HTTP_ERROR",
            f"HTTP {status} fetching URL.",
            detail=f"{url}: {e}",
            hint="The page may be behind a login wall, blocked, or deleted. Try lean_web_search for an alternative source.",
        )
    except httpx.TimeoutException as e:
        return error_response(
            "SCRAPE_TIMEOUT",
            "Request timed out.",
            detail=f"{url}: {e}",
            hint="The server is slow or unreachable. Try again or use lean_web_search for a similar page.",
        )
    except httpx.RequestError as e:
        return error_response(
            "SCRAPE_NETWORK_ERROR",
            "Failed to communicate with the server.",
            detail=f"{url}: {e}",
            hint="Check host connectivity, or try a different URL.",
        )

    content_type = response.headers.get("Content-Type", "").split(";")[0].strip().lower()
    is_pdf = looks_like_pdf or content_type == "application/pdf"

    if is_pdf:
        pdf_filename = re.sub(r"[^\w.\-]", "_", stripped_url.rsplit("/", 1)[-1] or "downloaded.pdf")
        if not pdf_filename.lower().endswith(".pdf"):
            pdf_filename += ".pdf"
        pdf_path = CACHE_DIR / pdf_filename
        pdf_path.write_bytes(response.content)
        return success_response(
            {"cached_path": str(pdf_path), "url": url},
            message="URL is a PDF; downloaded to cache. Use lean_pdf_read_text or lean_pdf_read_outline on the cached_path above.",
        )

    if not (content_type.startswith("text/html") or content_type.startswith("application/xhtml")):
        return error_response(
            "SCRAPE_UNSUPPORTED_CONTENT_TYPE",
            f"URL did not return an HTML page (Content-Type: {content_type or 'unknown'}).",
            detail=url,
            hint="This tool extracts article/documentation body text from HTML pages only.",
        )

    try:
        import trafilatura

        body_md = trafilatura.extract(
            response.text,
            output_format="markdown",
            favor_precision=favor_precision,
            include_links=include_links,
            include_images=False,
            include_tables=True,
        )
    except Exception:
        body_md = None

    if not body_md or not body_md.strip():
        return error_response(
            "SCRAPE_EMPTY",
            "No extractable body content found.",
            detail=url,
            hint="The page may be a JavaScript SPA or behind a paywall. Try a different URL or use lean_web_search.",
        )

    doc_meta = None
    try:
        import trafilatura

        doc_meta = trafilatura.extract_metadata(response.text)
    except Exception:
        doc_meta = None

    title = getattr(doc_meta, "title", None) if doc_meta else None
    sitename = getattr(doc_meta, "sitename", None) if doc_meta else None
    date = getattr(doc_meta, "date", None) if doc_meta else None

    if not include_links:
        body_md = _strip_markdown_links(body_md)

    blocks = _split_markdown_blocks(body_md)
    cap = max_chars if max_chars is not None else SCRAPE_MAX_CHARS_DEFAULT

    base_meta = {
        "url": url,
        "final_url": str(response.url),
        "title": title,
        "sitename": sitename,
        "date": date,
        "content_type": content_type,
    }

    if query and query.strip():
        result = _filter_by_query(blocks, query, offset=offset)
        capped_text, n_used = _cap_blocks_by_chars(result.items, cap)
        char_truncated = n_used < len(result.items)
        message = (
            f"No direct matches for query '{query}'. Showing top section." if result.no_match else None
        )
        meta = dict(base_meta)
        meta.update(
            {
                "filtered_by_query": query,
                "total_matches": result.total_matches,
                "offset": offset,
                "char_count_returned": len(capped_text),
            }
        )
        if result.next_offset is not None:
            meta["next_offset"] = result.next_offset
        elif char_truncated:
            meta["next_offset"] = offset + n_used
        return success_response(
            capped_text,
            message=message,
            truncated=(result.truncated or char_truncated),
            metadata=meta,
        )

    remaining_blocks = blocks[offset:]
    returned_text, n_used = _cap_blocks_by_chars(remaining_blocks, cap)
    end_idx = offset + n_used
    has_more = end_idx < len(blocks)
    full_text = "\n\n".join(blocks)

    meta = dict(base_meta)
    meta.update(
        {
            "char_count_returned": len(returned_text),
            "char_count_total": len(full_text),
            "offset": offset,
        }
    )
    if has_more:
        meta["next_offset"] = end_idx

    return success_response(returned_text, truncated=has_more, metadata=meta)


# ===========================================================================
# Excel Tools (ported bug-for-bug from lean_mcp.py, except the row cap below
# — an explicit deliberate deviation: the plan's largest unbounded-context
# hole, undercutting the plugin's whole premise)
# ===========================================================================
@mcp.tool(name="lean_excel_inspect")
def excel_inspect(path: str) -> str:
    """Returns sheet names, dimensions, and header row for an Excel workbook."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return error_response("XLSX_NOT_FOUND", "Spreadsheet does not exist", str(resolved))

    import openpyxl

    try:
        wb = openpyxl.load_workbook(resolved)
        active_sheet = wb.sheetnames[0]
        ws = wb[active_sheet]
        first_rows = list(ws.iter_rows(max_row=1, values_only=True))
        headers = list(first_rows[0]) if first_rows else []
        meta = {
            "sheet_names": wb.sheetnames,
            "active_sheet": active_sheet,
            "headers": headers,
            "dimensions": ws.dimensions,
            "max_rows": ws.max_row,
            "max_cols": ws.max_column,
        }
        wb.close()
        return success_response(meta)
    except Exception as e:
        return error_response("XLSX_CORRUPT", f"Cannot open workbook: {e}", str(resolved))


@mcp.tool(name="lean_excel_read_rows")
def excel_read_rows(
    path: str,
    sheet: Optional[str] = None,
    range_box: Optional[str] = None,
    output_format: Literal["json", "csv"] = "json",
    query: Optional[str] = None,
    offset: int = 0,
) -> str:
    """Reads rows from an Excel file as structured JSON (or CSV). Supports query filtering and a row cap."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return error_response("XLSX_NOT_FOUND", "Spreadsheet does not exist", str(resolved))

    import openpyxl

    try:
        wb = openpyxl.load_workbook(resolved)
    except Exception as e:
        return error_response("XLSX_CORRUPT", f"Cannot open workbook: {e}", str(resolved))

    ws_name = sheet or wb.sheetnames[0]
    if ws_name not in wb.sheetnames:
        wb.close()
        return error_response("XLSX_BAD_SHEET", f"Sheet '{ws_name}' not found", str(resolved))

    ws = wb[ws_name]
    iter_kwargs: Dict[str, Any] = {"values_only": True}
    if range_box:
        try:
            min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(range_box)
            iter_kwargs.update(
                min_row=min_row,
                max_row=min(max_row, ws.max_row or 1),
                min_col=min_col,
                max_col=min(max_col, ws.max_column or 1),
            )
        except Exception as e:
            wb.close()
            return error_response("XLSX_BAD_RANGE", f"Invalid range '{range_box}': {e}", str(resolved))

    raw_rows = list(ws.iter_rows(**iter_kwargs))
    wb.close()

    if not raw_rows:
        return success_response([])

    headers = [str(c) if c is not None else f"col_{i+1}" for i, c in enumerate(raw_rows[0])]
    dict_rows = []
    for row in raw_rows[1:]:
        dict_rows.append(
            {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        )

    if query and query.strip():
        row_strings = [json.dumps(r) for r in dict_rows]
        result = _filter_by_query(row_strings, query, offset=offset)
        filtered_dicts = []
        for s in result.items:
            try:
                filtered_dicts.append(json.loads(s))
            except Exception:
                pass
        message = (
            f"No direct matches for query '{query}'. Showing top section." if result.no_match else None
        )
        meta = {"filtered_by_query": query, "total_matches": result.total_matches, "offset": offset}
        if result.next_offset is not None:
            meta["next_offset"] = result.next_offset
        return success_response(
            filtered_dicts or result.items, message=message, truncated=result.truncated, metadata=meta
        )

    # Deliberate fix (not a bug-for-bug port): the original tool had no cap
    # at all on `dict_rows`, unlike every other read tool in this codebase —
    # the largest unbounded-context hole in the whole plugin set. Page it
    # the same way as the query branch: `offset` + a fixed page size, with a
    # `truncated`/`next_offset` pair the model can follow.
    total_rows = len(dict_rows)
    page = dict_rows[offset : offset + EXCEL_MAX_ROWS_DEFAULT]
    has_more = offset + len(page) < total_rows
    row_meta = {"total_rows": total_rows, "offset": offset}
    if has_more:
        row_meta["next_offset"] = offset + len(page)

    if output_format == "json":
        return success_response(page, truncated=has_more, metadata=row_meta)
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in page:
            writer.writerow([r.get(h, "") for h in headers])
        return success_response(output.getvalue(), truncated=has_more, metadata=row_meta)


@mcp.tool(name="lean_excel_write_rows")
def excel_write_rows(
    path: str,
    csv_data: str,
    sheet: Optional[str] = None,
    dry_run: bool = False,
) -> str:
    """Writes CSV-formatted rows into an Excel sheet, creating the workbook/sheet if needed."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return error_response("XLSX_NOT_FOUND", "Spreadsheet does not exist", str(resolved))

    import openpyxl

    try:
        wb = openpyxl.load_workbook(resolved)
    except Exception as e:
        return error_response("XLSX_CORRUPT", f"Cannot open workbook: {e}", str(resolved))

    ws_name = sheet or wb.sheetnames[0]
    if ws_name not in wb.sheetnames:
        wb.close()
        return error_response("XLSX_BAD_SHEET", f"Sheet '{ws_name}' not found", str(resolved))

    ws = wb[ws_name]
    reader = csv.reader(io.StringIO(csv_data.strip()))
    parsed_rows = list(reader)

    if dry_run:
        wb.close()
        return success_response(
            {"rows_to_write": len(parsed_rows)}, message="[DRY RUN] Would write rows."
        )

    for row in parsed_rows:
        ws.append(row)

    wb.save(resolved)
    wb.close()
    return success_response(
        {"rows_appended": len(parsed_rows), "path": str(resolved)},
        message="Successfully appended rows.",
    )


# ===========================================================================
# PDF Tools (ported bug-for-bug from lean_mcp.py)
# ===========================================================================
@mcp.tool(name="lean_pdf_read_text")
def pdf_read_text(
    path: str,
    start_page: int = 1,
    end_page: Optional[int] = None,
    query: Optional[str] = None,
    offset: int = 0,
) -> str:
    """Reads PDF text with PyMuPDF layout analysis. Supports page ranges and query filtering."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return error_response("PDF_NOT_FOUND", "PDF does not exist", str(resolved))

    try:
        doc = fitz.open(str(resolved))
    except Exception as e:
        return error_response("PDF_CORRUPT", f"Cannot parse PDF: {e}", str(resolved))

    if doc.is_encrypted:
        doc.close()
        return error_response("PDF_ENCRYPTED", "PDF is password protected", str(resolved))

    total_pages = len(doc)
    s_page = max(1, start_page)
    e_page = min(total_pages, end_page) if end_page else total_pages

    extracted_pages = []
    for pno in range(s_page - 1, e_page):
        page = doc[pno]
        # `get_text("markdown")` is not a valid PyMuPDF text-extraction
        # option (valid ones: text, words, blocks, html, dict, json, rawdict,
        # xhtml, xml) — it raises an AssertionError rather than falling
        # through to the `or page.get_text()` the original code relied on,
        # which never actually executes. Fall back to plain text explicitly.
        try:
            text = page.get_text("markdown")
        except (AssertionError, ValueError):
            text = None
        text = text or page.get_text()
        extracted_pages.append(f"--- Page {pno + 1} ---\n" + text.strip())

    doc.close()

    if query and query.strip():
        result = _filter_by_query(extracted_pages, query, offset=offset)
        message = (
            f"No direct matches for query '{query}'. Showing top section." if result.no_match else None
        )
        meta = {
            "start_page": s_page,
            "end_page": e_page,
            "filtered_by_query": query,
            "total_matches": result.total_matches,
            "offset": offset,
        }
        if result.next_offset is not None:
            meta["next_offset"] = result.next_offset
        return success_response(result.items, message=message, truncated=result.truncated, metadata=meta)

    return success_response(
        extracted_pages,
        metadata={"start_page": s_page, "end_page": e_page, "total_pages": total_pages},
    )


@mcp.tool(name="lean_pdf_read_outline")
def pdf_read_outline(path: str) -> str:
    """Returns the table-of-contents/bookmark outline of a PDF, if it has one."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        return error_response("PDF_NOT_FOUND", "PDF does not exist", str(resolved))

    try:
        doc = fitz.open(str(resolved))
    except Exception as e:
        return error_response("PDF_CORRUPT", f"Cannot parse PDF: {e}", str(resolved))

    if doc.is_encrypted:
        doc.close()
        return error_response("PDF_ENCRYPTED", "PDF is password protected", str(resolved))

    toc = doc.get_toc()
    outline = [{"level": item[0], "title": item[1], "page": item[2]} for item in toc]
    doc.close()
    return success_response(outline)


def main() -> None:
    mcp.run()


if __name__ == "__main__":
    main()
